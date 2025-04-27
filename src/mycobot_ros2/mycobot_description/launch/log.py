#! /usr/bin/env python3

from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill, Font


class CSVToExcel:
    def __init__(self, files_with_headings, files_with_values, output_excel):
        self.files_with_headings = files_with_headings
        self.files_with_values = files_with_values
        self.output_excel = output_excel
        self.summary_data = {'Sheet1': [], 'Sheet2': [], 'Sheet3': []}
        self.bold_cells = []

    def combine_csv(self, heading_file, value_file):
        """Combine the two CSV files vertically."""
        df_heading = pd.read_csv(heading_file)
        df_values = pd.read_csv(value_file)
        combined_df = pd.concat([df_heading, df_values], ignore_index=True)
        return combined_df

    def prepare_summary(self, combined_df, sheet_name):
        """Prepare summary for the heading comparison."""
        summary_text = ""
        for head in combined_df.columns:
            try:
                value1 = pd.to_numeric(combined_df.iloc[0][head])
                value2 = pd.to_numeric(combined_df.iloc[1][head])
                difference = value1 - value2
                diff_str = str(difference)

                # Create the summary text and locate the position of the difference
                start_pos = len(f"{head} differs by ")
                end_pos = start_pos + len(diff_str)
                summary_text += f"{head} differs by {diff_str} "

                # Track position for bold formatting (start and end of the difference part)
                self.bold_cells.append((2, len(self.summary_data) + 1, start_pos, end_pos))

            except Exception as e:
                summary_text += f"{head} not numeric "

        self.summary_data[sheet_name].append(summary_text)

    def create_excel_file(self):
        """Create an Excel file with 3 sheets and summary."""
        with pd.ExcelWriter(self.output_excel, engine='openpyxl') as writer:
            for idx, (heading_file, value_file) in enumerate(zip(self.files_with_headings, self.files_with_values)):
                # Combine data from both CSV files
                combined_df = self.combine_csv(heading_file, value_file)

                # Write combined data to Excel
                combined_df.to_excel(writer, sheet_name=f'Sheet{idx + 1}', index=False)

                # Prepare the summary info
                sheet_name = f'Sheet{idx + 1}'
                self.prepare_summary(combined_df, sheet_name)

            # Create the Summary sheet
            summary_df = pd.DataFrame(self.summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

    def color_and_bold_cells(self):
        """Apply yellow fill and bold formatting to specific cells."""
        wb = load_workbook(self.output_excel)

        # Define Yellow Fill
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Color value cells in Sheet1/2/3
        for sheet_name in ['Sheet1', 'Sheet2', 'Sheet3']:
            ws = wb[sheet_name]

            # Row 2 and 3 → original two data rows (row 1 = header)
            for row in [2, 3]:
                for col in range(2, ws.max_column + 1):  # Start from column 2 to max column
                    cell = ws.cell(row=row, column=col)
                    cell.fill = yellow_fill

        # Bold specific cells in Summary sheet (difference part only)
        ws_summary = wb['Summary']
        for row, col, start, end in self.bold_cells:
            # Bold just the difference value
            cell = ws_summary.cell(row=row, column=col)
            cell_text = cell.value
            if cell_text:
                # Bold only the difference part (the numeric part of the difference)
                bold_font = Font(bold=True)
                # Extract the number part of the string
                num_part = cell_text[start:end]
                cell.value = cell_text[:start] + num_part + cell_text[end:]
                cell.font = bold_font

        # Save changes
        wb.save(self.output_excel)
        print(f"✅ Excel file '{self.output_excel}' created with colored values and bolded difference in Summary.")




# Example usage
files_with_headings = ['f1/file1.csv', 'f1/file2.csv', 'f1/file3.csv']
files_with_values = ['f2/file1.csv', 'f2/file2.csv', 'f2/file3.csv']
output_excel = 'final_output3.xlsx'

# Instantiate the class and call methods
csv_to_excel = CSVToExcel(files_with_headings, files_with_values, output_excel)
csv_to_excel.create_excel_file()
csv_to_excel.color_and_bold_cells()

